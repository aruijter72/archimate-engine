#!/usr/bin/env python3
"""Generate/update model.archimate from model-input.xlsx.

Usage:
    python scripts/generate_model.py [--excel PATH] [--output PATH]
"""

import argparse
import sys
import uuid
from xml.dom import minidom
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent

XSI = 'http://www.w3.org/2001/XMLSchema-instance'
ARCH = 'http://www.archimatetool.com/archimate'

ET.register_namespace('xsi', XSI)
ET.register_namespace('archimate', ARCH)


def new_id():
    return 'id-' + uuid.uuid4().hex[:12]


def read_sheet(wb, sheet_name, required_cols):
    """Read a sheet and return list of dicts. Skips rows where required_cols are empty."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [str(c.value or '').strip().rstrip(' *') for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        record = {headers[i]: str(v).strip() if v is not None else '' for i, v in enumerate(raw) if i < len(headers)}
        if any(not record.get(col) for col in required_cols):
            continue
        rows.append(record)
    return rows


def read_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    business = read_sheet(wb, 'Business', ['Name', 'Type'])
    application = read_sheet(wb, 'Application', ['Name', 'Type'])
    technology = read_sheet(wb, 'Technology', ['Name', 'Type'])
    relationships = read_sheet(wb, 'Relationships', ['Source ID', 'Target ID', 'Type'])
    views = read_sheet(wb, 'Views', ['View Name'])
    return business, application, technology, relationships, views


def add_folder(root, name, ftype):
    f = ET.SubElement(root, 'folder')
    f.set('name', name)
    f.set('id', new_id())
    f.set('type', ftype)
    return f


def add_doc(parent, text):
    if text:
        d = ET.SubElement(parent, 'documentation')
        d.text = text


def add_prop(parent, key, value):
    if value:
        p = ET.SubElement(parent, 'property')
        p.set('key', key)
        p.set('value', value)


def add_elements(folder, items, layer, element_map):
    for item in items:
        name = item['Name'].strip()
        etype = item['Type'].strip()
        if not name or not etype:
            continue
        eid = item.get('ID', '').strip() or new_id()
        item['_id'] = eid
        element_map[eid] = {'layer': layer, 'item': item}

        el = ET.SubElement(folder, 'element')
        el.set(f'{{{XSI}}}type', f'archimate:{etype}')
        el.set('name', name)
        el.set('id', eid)
        add_doc(el, item.get('Description', '') or item.get('Documentation', ''))
        add_prop(el, 'Tags', item.get('Tags', ''))
        add_prop(el, 'Status', item.get('Status', ''))


def layout_view(diagram, included_ids, element_map, rel_map):
    """Place DiagramObjects in a layered grid; add Connections."""
    ELEM_W, ELEM_H = 130, 55
    PAD_X, PAD_Y = 20, 20
    LAYER_GAP = 50
    PER_ROW = 5

    layer_order = ['business', 'application', 'technology', 'other']
    buckets = {l: [] for l in layer_order}
    for eid in included_ids:
        if eid in element_map:
            layer = element_map[eid]['layer']
            buckets.setdefault(layer, []).append(eid)

    diag_obj_map = {}  # element_id -> diag_object_id
    y_offset = PAD_Y

    for layer in layer_order:
        eids = buckets.get(layer, [])
        if not eids:
            continue
        for i, eid in enumerate(eids):
            col = i % PER_ROW
            row_n = i // PER_ROW
            x = PAD_X + col * (ELEM_W + PAD_X)
            y = y_offset + row_n * (ELEM_H + PAD_Y)

            dobj_id = new_id()
            diag_obj_map[eid] = dobj_id

            child = ET.SubElement(diagram, 'child')
            child.set(f'{{{XSI}}}type', 'archimate:DiagramObject')
            child.set('id', dobj_id)
            child.set('archimateElement', eid)

            bounds = ET.SubElement(child, 'bounds')
            bounds.set('x', str(x))
            bounds.set('y', str(y))
            bounds.set('width', str(ELEM_W))
            bounds.set('height', str(ELEM_H))

        n_rows = max(1, (len(eids) + PER_ROW - 1) // PER_ROW)
        y_offset += n_rows * (ELEM_H + PAD_Y) + LAYER_GAP

    for rid, rel in rel_map.items():
        src = rel['source']
        tgt = rel['target']
        if src in diag_obj_map and tgt in diag_obj_map:
            conn = ET.SubElement(diagram, 'child')
            conn.set(f'{{{XSI}}}type', 'archimate:Connection')
            conn.set('id', new_id())
            conn.set('source', diag_obj_map[src])
            conn.set('target', diag_obj_map[tgt])
            conn.set('archimateRelationship', rid)


def build_model(business, application, technology, relationships, views):
    root = ET.Element(f'{{{ARCH}}}model')
    root.set('name', 'ArchiMate Model')
    root.set('id', new_id())
    root.set('version', '4.0')

    folders = {}
    for name, ftype in [
        ('Strategy', 'strategy'),
        ('Business', 'business'),
        ('Application', 'application'),
        ('Technology & Physical', 'technology'),
        ('Motivation', 'motivation'),
        ('Implementation & Migration', 'implementation_migration'),
        ('Other', 'other'),
        ('Relations', 'relations'),
        ('Views', 'diagrams'),
    ]:
        folders[ftype] = add_folder(root, name, ftype)

    element_map = {}
    add_elements(folders['business'], business, 'business', element_map)
    add_elements(folders['application'], application, 'application', element_map)
    add_elements(folders['technology'], technology, 'technology', element_map)

    rel_map = {}
    for rel in relationships:
        src_id = rel.get('Source ID', '').strip()
        tgt_id = rel.get('Target ID', '').strip()
        rtype = rel.get('Type', '').strip()
        # Skip placeholder rows from the sample template
        if not src_id or not tgt_id or not rtype or src_id.startswith('<'):
            continue
        # Validate IDs exist in element map
        if src_id not in element_map or tgt_id not in element_map:
            print(f'  Warning: skipping relationship {src_id} -> {tgt_id} (unknown element ID)')
            continue
        rid = rel.get('ID', '').strip() or new_id()
        rel['_id'] = rid
        rel_map[rid] = {'source': src_id, 'target': tgt_id, 'type': rtype,
                        'name': rel.get('Name', ''), 'accessType': rel.get('Access Type', '')}

        el = ET.SubElement(folders['relations'], 'element')
        el.set(f'{{{XSI}}}type', f'archimate:{rtype}')
        el.set('id', rid)
        el.set('source', src_id)
        el.set('target', tgt_id)
        if rel.get('Name'):
            el.set('name', rel['Name'])
        if rel.get('Access Type'):
            el.set('accessType', rel['Access Type'].upper())

    if not views:
        views = [{'View Name': 'Default View', 'Include Elements': 'ALL',
                  'Description': 'Auto-generated default view', 'Default View': 'Yes'}]

    all_element_ids = list(element_map.keys())

    for view in views:
        vname = view.get('View Name', '').strip()
        if not vname:
            continue
        vid = view.get('View ID', '').strip() or new_id()

        diag = ET.SubElement(folders['diagrams'], 'element')
        diag.set(f'{{{XSI}}}type', 'archimate:ArchimateDiagramModel')
        diag.set('name', vname)
        diag.set('id', vid)
        add_doc(diag, view.get('Description', ''))

        include_str = view.get('Include Elements', 'ALL').strip()
        if not include_str or include_str.upper() == 'ALL':
            included = all_element_ids
        else:
            included = [i.strip() for i in include_str.split(',') if i.strip() and not i.strip().startswith('<')]
            # Keep only valid IDs
            included = [i for i in included if i in element_map]

        layout_view(diag, included, element_map, rel_map)

    return root


def prettify(root):
    raw = ET.tostring(root, encoding='unicode')
    dom = minidom.parseString(raw)
    pretty = dom.toprettyxml(indent='  ', encoding='UTF-8')
    return pretty.decode('utf-8')


def main():
    parser = argparse.ArgumentParser(description='Generate model.archimate from Excel input.')
    parser.add_argument('--excel', default=str(BASE_DIR / 'model-input.xlsx'))
    parser.add_argument('--output', default=str(BASE_DIR / 'model.archimate'))
    args = parser.parse_args()

    excel_path = Path(args.excel)
    output_path = Path(args.output)

    if not excel_path.exists():
        print(f'Error: {excel_path} not found. Run scripts/create_template.py first.')
        sys.exit(1)

    print(f'Reading {excel_path} ...')
    business, application, technology, relationships, views = read_excel(excel_path)
    print(f'  Business: {len(business)}  Application: {len(application)}  '
          f'Technology: {len(technology)}  Relationships: {len(relationships)}  Views: {len(views)}')

    print('Building model.archimate ...')
    root = build_model(business, application, technology, relationships, views)

    output_path.write_text(prettify(root), encoding='utf-8')
    print(f'Saved: {output_path}')


if __name__ == '__main__':
    main()

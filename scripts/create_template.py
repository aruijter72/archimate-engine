#!/usr/bin/env python3
"""Create the ArchiMate model input Excel template with sample data."""

import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BUSINESS_TYPES = [
    'BusinessActor', 'BusinessRole', 'BusinessCollaboration',
    'BusinessInterface', 'BusinessProcess', 'BusinessFunction',
    'BusinessInteraction', 'BusinessEvent', 'BusinessService',
    'BusinessObject', 'Contract', 'Product',
]
APPLICATION_TYPES = [
    'ApplicationComponent', 'ApplicationCollaboration',
    'ApplicationInterface', 'ApplicationFunction',
    'ApplicationInteraction', 'ApplicationProcess',
    'ApplicationEvent', 'ApplicationService', 'DataObject',
]
TECHNOLOGY_TYPES = [
    'Node', 'Device', 'SystemSoftware', 'TechnologyCollaboration',
    'TechnologyInterface', 'Path', 'CommunicationNetwork',
    'TechnologyFunction', 'TechnologyProcess', 'TechnologyInteraction',
    'TechnologyEvent', 'TechnologyService', 'Artifact',
]
RELATIONSHIP_TYPES = [
    'AssignmentRelationship', 'RealizationRelationship',
    'ServingRelationship', 'AccessRelationship',
    'FlowRelationship', 'TriggeringRelationship',
    'CompositionRelationship', 'AggregationRelationship',
    'AssociationRelationship', 'InfluenceRelationship',
    'SpecializationRelationship',
]

HEADER_BG = '1F3864'
HEADER_FG = 'FFFFFF'
LAYER_BG = {
    'business': 'FFF2CC',
    'application': 'DAE8FC',
    'technology': 'D5E8D4',
    'relations': 'F8CECC',
    'views': 'E1D5E7',
}
TAB_COLORS = {
    'business': 'FFD700',
    'application': '4472C4',
    'technology': '70AD47',
    'relations': 'FF0000',
    'views': '7030A0',
    'instructions': '808080',
}


def hdr_font():
    return Font(bold=True, color=HEADER_FG, name='Arial', size=10)


def hdr_fill():
    return PatternFill('solid', fgColor=HEADER_BG)


def hdr_align():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)


def layer_fill(key):
    return PatternFill('solid', fgColor=LAYER_BG[key])


def apply_header(cell, value):
    cell.value = value
    cell.font = hdr_font()
    cell.fill = hdr_fill()
    cell.alignment = hdr_align()


def set_data_row(ws, row, layer_key):
    fill = layer_fill(layer_key)
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = Font(name='Arial', size=10)
        c.alignment = Alignment(vertical='top', wrap_text=(col in [4, 5]))


def add_instructions(wb):
    ws = wb.active
    ws.title = 'Instructions'
    ws.sheet_properties.tabColor = TAB_COLORS['instructions']

    ws.merge_cells('A1:H1')
    t = ws['A1']
    t.value = 'ArchiMate Model Input Form'
    t.font = Font(bold=True, size=20, name='Arial', color=HEADER_FG)
    t.fill = PatternFill('solid', fgColor=HEADER_BG)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 46

    ws.merge_cells('A2:H2')
    s = ws['A2']
    s.value = 'Repository: https://github.com/aruijter72/archimate-engine  |  Fill in each layer tab, then push to GitHub'
    s.font = Font(size=10, name='Arial', color='555555')
    s.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    sections = [
        ('HOW TO USE THIS FORM', None),
        ('Step 1', 'Open the Business, Application, and Technology tabs and add your architecture elements. Leave the ID column blank — IDs are auto-assigned by the generator.'),
        ('Step 2', 'Open the Relationships tab and link elements by entering their IDs in Source ID / Target ID. Choose the ArchiMate relationship type from the dropdown.'),
        ('Step 3', 'Open the Views tab and define one or more named views. Enter ALL to include every element, or list specific element IDs separated by commas.'),
        ('Step 4', 'Save the file. Commit and push to GitHub (main branch). The GitHub Actions pipeline auto-generates model.archimate and docs/index.html, then publishes to GitHub Pages.'),
        ('Step 5', 'Claude can also update this form or model.archimate directly from a chat prompt, then commit the result to GitHub.'),
        ('COLUMN GUIDE — ELEMENT SHEETS', None),
        ('ID', 'Auto-generated UUID. Leave blank when adding new rows. Do not edit existing IDs.'),
        ('Name *', 'Required. The element name as it appears in the diagram.'),
        ('Type *', 'Required. Select from the dropdown of valid ArchiMate types for this layer.'),
        ('Description', 'Short text shown as a tooltip in the HTML viewer.'),
        ('Documentation', 'Extended notes stored in the model and shown in the details panel.'),
        ('Tags', 'Comma-separated tags for filtering, e.g. "core, legacy, customer-facing".'),
        ('Status', 'Lifecycle stage: Active, Draft, Deprecated, or To Be.'),
        ('COLUMN GUIDE — RELATIONSHIPS', None),
        ('Source ID / Target ID *', 'Copy-paste the ID from the element sheets. Both fields are required.'),
        ('Type *', 'Select the ArchiMate relationship type. See ArchiMate 3 spec for semantics.'),
        ('Name', 'Optional label shown on the relationship arrow in the diagram.'),
        ('Access Type', 'Only for AccessRelationship: Read, Write, ReadWrite, or Access.'),
        ('COLUMN GUIDE — VIEWS', None),
        ('View Name *', 'Required. A descriptive name for the diagram.'),
        ('Include Elements', 'Enter ALL to show everything, or comma-separated element IDs for a scoped view.'),
        ('Default View', 'Set to Yes for the view opened by default in the HTML viewer.'),
    ]

    row = 4
    for label, desc in sections:
        ws.row_dimensions[row].height = 24
        if desc is None:
            ws.merge_cells(f'A{row}:H{row}')
            c = ws.cell(row=row, column=1, value=label)
            c.font = Font(bold=True, size=11, name='Arial', color=HEADER_FG)
            c.fill = PatternFill('solid', fgColor='2F5496')
            c.alignment = Alignment(horizontal='left', vertical='center')
        else:
            lc = ws.cell(row=row, column=1, value=label)
            dc = ws.cell(row=row, column=2, value=desc)
            lc.font = Font(bold=True, size=10, name='Arial', color='1F3864')
            dc.font = Font(size=10, name='Arial')
            dc.alignment = Alignment(wrap_text=True, vertical='top')
            lc.alignment = Alignment(vertical='top')
            if row % 2 == 0:
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor='F7F7F7')
        row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 80
    return ws


def add_element_sheet(wb, name, layer_key, element_types, samples):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = TAB_COLORS[layer_key]

    headers = ['ID', 'Name *', 'Type *', 'Description', 'Documentation', 'Tags', 'Status']
    for col, h in enumerate(headers, 1):
        apply_header(ws.cell(row=1, column=col), h)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    widths = [24, 32, 28, 44, 56, 28, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Type dropdown
    dv_type = DataValidation(
        type='list', formula1='"' + ','.join(element_types) + '"',
        allow_blank=False, showErrorMessage=True,
        errorTitle='Invalid Type', error='Select a valid type from the list.',
    )
    ws.add_data_validation(dv_type)
    dv_type.add('C2:C2000')

    # Status dropdown
    dv_status = DataValidation(
        type='list', formula1='"Active,Draft,Deprecated,To Be"', allow_blank=True,
    )
    ws.add_data_validation(dv_status)
    dv_status.add('G2:G2000')

    # Sample data
    for i, row_data in enumerate(samples, 2):
        ws.cell(row=i, column=1).value = ''  # ID blank — auto-generated
        ws.cell(row=i, column=2).value = row_data[0]
        ws.cell(row=i, column=3).value = row_data[1]
        ws.cell(row=i, column=4).value = row_data[2]
        ws.cell(row=i, column=5).value = ''
        ws.cell(row=i, column=6).value = row_data[3] if len(row_data) > 3 else ''
        ws.cell(row=i, column=7).value = row_data[4] if len(row_data) > 4 else 'Active'
        set_data_row(ws, i, layer_key)

    # Empty rows for new entries
    for row in range(len(samples) + 2, len(samples) + 22):
        set_data_row(ws, row, layer_key)

    return ws


def add_relationships_sheet(wb, samples):
    ws = wb.create_sheet('Relationships')
    ws.sheet_properties.tabColor = TAB_COLORS['relations']

    headers = ['ID', 'Source ID *', 'Source Name', 'Target ID *', 'Target Name',
               'Type *', 'Name', 'Documentation', 'Access Type']
    for col, h in enumerate(headers, 1):
        apply_header(ws.cell(row=1, column=col), h)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    widths = [24, 24, 30, 24, 30, 28, 28, 50, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    dv_rel = DataValidation(
        type='list', formula1='"' + ','.join(RELATIONSHIP_TYPES) + '"',
        allow_blank=False, showErrorMessage=True,
    )
    ws.add_data_validation(dv_rel)
    dv_rel.add('F2:F2000')

    dv_access = DataValidation(
        type='list', formula1='"Read,Write,ReadWrite,Access"', allow_blank=True,
    )
    ws.add_data_validation(dv_access)
    dv_access.add('I2:I2000')

    for i, row_data in enumerate(samples, 2):
        ws.cell(row=i, column=1).value = ''
        ws.cell(row=i, column=2).value = row_data[0]  # Source ID placeholder
        ws.cell(row=i, column=3).value = row_data[1]  # Source name hint
        ws.cell(row=i, column=4).value = row_data[2]  # Target ID placeholder
        ws.cell(row=i, column=5).value = row_data[3]  # Target name hint
        ws.cell(row=i, column=6).value = row_data[4]  # Type
        ws.cell(row=i, column=7).value = row_data[5] if len(row_data) > 5 else ''
        set_data_row(ws, i, 'relations')

    for row in range(len(samples) + 2, len(samples) + 22):
        set_data_row(ws, row, 'relations')

    return ws


def add_views_sheet(wb, samples):
    ws = wb.create_sheet('Views')
    ws.sheet_properties.tabColor = TAB_COLORS['views']

    headers = ['View ID', 'View Name *', 'Description',
               'Documentation', 'Include Elements (IDs or ALL)', 'View Type', 'Default View']
    for col, h in enumerate(headers, 1):
        apply_header(ws.cell(row=1, column=col), h)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    widths = [24, 32, 44, 56, 50, 22, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    dv_vtype = DataValidation(
        type='list', formula1='"ArchiMate View,Layered View,Map View"', allow_blank=True,
    )
    ws.add_data_validation(dv_vtype)
    dv_vtype.add('F2:F2000')

    dv_default = DataValidation(
        type='list', formula1='"Yes,No"', allow_blank=True,
    )
    ws.add_data_validation(dv_default)
    dv_default.add('G2:G2000')

    for i, row_data in enumerate(samples, 2):
        ws.cell(row=i, column=1).value = ''
        ws.cell(row=i, column=2).value = row_data[0]
        ws.cell(row=i, column=3).value = row_data[1]
        ws.cell(row=i, column=4).value = ''
        ws.cell(row=i, column=5).value = row_data[2]
        ws.cell(row=i, column=6).value = 'ArchiMate View'
        ws.cell(row=i, column=7).value = row_data[3]
        set_data_row(ws, i, 'views')

    for row in range(len(samples) + 2, len(samples) + 8):
        set_data_row(ws, row, 'views')

    return ws


def create_template(output_path):
    wb = Workbook()
    add_instructions(wb)

    business_samples = [
        ('Customer', 'BusinessActor', 'End user interacting with the platform', 'core, external', 'Active'),
        ('Account Manager', 'BusinessRole', 'Internal role managing customer accounts', 'internal', 'Active'),
        ('Order Management', 'BusinessProcess', 'Process for handling customer orders end-to-end', 'core', 'Active'),
        ('Product Catalog Service', 'BusinessService', 'Provides product information to customers', 'core', 'Active'),
        ('Order', 'BusinessObject', 'Represents a customer order', 'data', 'Active'),
    ]
    add_element_sheet(wb, 'Business', 'business', BUSINESS_TYPES, business_samples)

    application_samples = [
        ('Order Management System', 'ApplicationComponent', 'Core application for order processing', 'core', 'Active'),
        ('CRM System', 'ApplicationComponent', 'Customer relationship management platform', 'core', 'Active'),
        ('API Gateway', 'ApplicationComponent', 'Central API entry point for all services', 'infrastructure', 'Active'),
        ('Order API', 'ApplicationInterface', 'REST API exposing order operations', 'api', 'Active'),
        ('Order Data', 'DataObject', 'Data entity representing an order record', 'data', 'Active'),
    ]
    add_element_sheet(wb, 'Application', 'application', APPLICATION_TYPES, application_samples)

    technology_samples = [
        ('Application Server', 'Node', 'Hosts core business applications', 'infrastructure', 'Active'),
        ('Database Server', 'Node', 'Relational database for transactional data', 'infrastructure', 'Active'),
        ('Kubernetes Cluster', 'SystemSoftware', 'Container orchestration platform', 'infrastructure', 'Active'),
        ('Internal Network', 'CommunicationNetwork', 'Internal corporate network', 'network', 'Active'),
        ('PostgreSQL', 'SystemSoftware', 'Primary relational database engine', 'database', 'Active'),
    ]
    add_element_sheet(wb, 'Technology', 'technology', TECHNOLOGY_TYPES, technology_samples)

    # Relationships use placeholder text — IDs filled after generation
    rel_samples = [
        ('<Business Actor ID>', 'Customer', '<Business Role ID>', 'Account Manager', 'AssociationRelationship'),
        ('<App Component ID>', 'Order Management System', '<Business Service ID>', 'Order Management', 'RealizationRelationship'),
        ('<App Component ID>', 'CRM System', '<Business Actor ID>', 'Customer', 'ServingRelationship'),
        ('<Tech Node ID>', 'Application Server', '<App Component ID>', 'Order Management System', 'AssignmentRelationship'),
        ('<Tech Node ID>', 'Database Server', '<App Component ID>', 'Order Management System', 'AssignmentRelationship'),
    ]
    add_relationships_sheet(wb, rel_samples)

    view_samples = [
        ('Full Architecture View', 'Complete view of all architecture layers', 'ALL', 'Yes'),
        ('Business Layer', 'Business processes, actors, and services only', '<comma-sep business IDs>', 'No'),
        ('Application Architecture', 'Application components and their connections', '<comma-sep app IDs>', 'No'),
    ]
    add_views_sheet(wb, view_samples)

    wb.save(output_path)
    print(f'Template created: {output_path}')


if __name__ == '__main__':
    output = sys.argv[1] if len(sys.argv) > 1 else 'model-input.xlsx'
    create_template(Path(output))
